#!/usr/bin/env python3
"""汇总 data/results/ 下所有评分/运行结果，按试卷分sheet输出 xlsx。

每道题下按模型列出：得分、首轮耗时、次轮耗时、首轮Token、次轮Token
"""

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
GRADES_DIR = ROOT / "data/results/grades"
RUNS_DIR = ROOT / "data/results/runs"

PAPERS = ("2026-national-i-math", "2026-national-ii-math")


# ---- style helpers ----
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
QID_FILL = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
QID_FONT = Font(bold=True, size=11, color="FFFFFF")
SEPARATOR_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")

COLS = ["题目", "模型", "得分", "满分", "首轮耗时(s)", "次轮耗时(s)", "首轮Token", "次轮Token"]
COL_WIDTHS = [8, 28, 8, 8, 13, 13, 12, 12]


def sorted_items(ids: list[str]) -> list[str]:
    def key(s: str):
        num = s.rsplit("q", 1)[-1]
        return int(num) if num.isdigit() else num
    return sorted(set(ids), key=key)


def parse_filename(stem: str) -> tuple[str, str] | None:
    for paper in PAPERS:
        if stem.endswith(paper):
            model_name = stem[: -len(paper) - 1]
            return model_name, paper
    return None


def build_tables():
    """返回 {paper: {model: {item_id: {score, max_score, first_lat, final_lat, first_tok, final_tok}}}}"""
    # 先读 grades
    grades_map: dict[str, dict[str, dict]] = defaultdict(lambda: defaultdict(dict))
    for f in sorted(GRADES_DIR.glob("*.jsonl")):
        result = parse_filename(f.stem)
        if result is None:
            continue
        model, paper = result
        with open(f) as fh:
            for line in fh:
                if not line.strip():
                    continue
                g = json.loads(line)
                grades_map[paper][model][g["item_id"]] = {
                    "score": g["score"], "max_score": g["max_score"],
                }

    # 再读 runs
    for f in sorted(RUNS_DIR.glob("*.jsonl")):
        result = parse_filename(f.stem)
        if result is None:
            continue
        model, paper = result
        with open(f) as fh:
            for line in fh:
                if not line.strip():
                    continue
                r = json.loads(line)
                rr = r.get("raw_responses", {})
                first_p = (rr.get("first") or {}).get("payload") or {}
                final_p = (rr.get("final") or {}).get("payload") or {}
                entry = grades_map[paper][model].setdefault(r["item_id"], {"score": "", "max_score": ""})
                entry["first_lat"] = (rr.get("first") or {}).get("latency_seconds")
                entry["final_lat"] = (rr.get("final") or {}).get("latency_seconds")
                entry["first_tok"] = first_p.get("usage", {}).get("total_tokens")
                entry["final_tok"] = final_p.get("usage", {}).get("total_tokens")
    return grades_map


def write_paper_sheet(wb, paper: str, data: dict[str, dict[str, dict]]):
    """一张试卷一个 sheet：先模型总分汇总，后逐题明细"""
    models = sorted(data)
    all_qids = sorted_items([qid for m in models for qid in data[m]])

    def model_total(m):
        return sum((data[m].get(q, {}).get("score") or 0) for q in all_qids)
    model_order = sorted(models, key=model_total, reverse=True)

    ws = wb.create_sheet(title=paper[:31])
    ws.freeze_panes = "A2"
    row = 1

    # ========== 第一部分：模型总分汇总 ==========
    cell = ws.cell(row=row, column=1, value=f"{paper} 模型总分汇总")
    cell.font = Font(bold=True, size=12)
    row += 1

    summary_headers = ["模型", "总分", "总首轮耗时(s)", "总次轮耗时(s)", "总首轮Token", "总次轮Token"]
    for ci, name in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER
    row += 1

    for model in model_order:
        total_score = sum((data[model].get(q, {}).get("score") or 0) for q in all_qids)
        total_f_lat = sum(v for q in all_qids if (v := data[model].get(q, {}).get("first_lat")) is not None)
        total_fl_lat = sum(v for q in all_qids if (v := data[model].get(q, {}).get("final_lat")) is not None)
        total_f_tok = sum(v for q in all_qids if (v := data[model].get(q, {}).get("first_tok")) is not None)
        total_fl_tok = sum(v for q in all_qids if (v := data[model].get(q, {}).get("final_tok")) is not None)
        vals = [model, total_score, round(total_f_lat, 1), round(total_fl_lat, 1), total_f_tok, total_fl_tok]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font, cell.alignment, cell.border = NORMAL_FONT, CENTER, THIN_BORDER
        row += 1

    # 空行分隔
    row += 2

    # ========== 第二部分：逐题明细 ==========
    # 明细列头
    for ci, name in enumerate(COLS, start=1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER
    row += 1

    for qid in all_qids:
        # 题目行
        cell = ws.cell(row=row, column=1, value=qid.split("-")[-1])
        cell.font, cell.fill, cell.alignment, cell.border = QID_FONT, QID_FILL, CENTER, THIN_BORDER
        for ci in range(2, len(COLS) + 1):
            cell = ws.cell(row=row, column=ci)
            cell.fill, cell.border = QID_FILL, THIN_BORDER
        row += 1

        for model in model_order:
            d = data.get(model, {}).get(qid, {})
            score = d.get("score")
            vals = [
                qid.split("-")[-1],
                model,
                round(score, 1) if isinstance(score, float) else (score if score != "" and score is not None else "-"),
                d.get("max_score", ""),
                round(d["first_lat"], 1) if d.get("first_lat") is not None else "-",
                round(d["final_lat"], 1) if d.get("final_lat") is not None else "-",
                d.get("first_tok") if d.get("first_tok") is not None else "-",
                d.get("final_tok") if d.get("final_tok") is not None else "-",
            ]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font, cell.alignment, cell.border = NORMAL_FONT, CENTER, THIN_BORDER
            row += 1

    # 列宽
    for ci, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main():
    data = build_tables()
    papers = sorted(data)
    print(f"发现 {len(papers)} 张试卷: {papers}")

    wb = Workbook()
    wb.remove(wb.active)

    for paper in papers:
        print(f"  生成 {paper} ...")
        write_paper_sheet(wb, paper, data[paper])

    output = ROOT / "grades_summary.xlsx"
    wb.save(output)
    print(f"\n已生成: {output}")


if __name__ == "__main__":
    main()
